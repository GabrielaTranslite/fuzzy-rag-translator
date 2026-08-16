"""
build_gold_set.py
Convert the hand-filled gold_worksheet.xlsx into eval/gold_set.jsonl.

Run:
    python eval/build_gold_set.py

Re-run it any time you edit the worksheet. It always reads the CURRENT file,
so you can keep improving your reference translations and just rebuild - nothing
is frozen. Save/close the xlsx in Excel first so your latest edits are on disk.

What it does:
  1. Reads the three sheets (fuzzy_real, edited, invented), skipping the header
     row and row 2 (the illustrative EXAMPLE).
  2. Builds one JSON record per case (fields depend on the category).
  3. Validates the data and prints a report - it never silently drops a row:
       - base_id (edited sheet) must exist in the translation memory
       - every $placeholder in `query` must also appear in `reference`
       - edit_type / requires_agreement / flag must use the allowed vocabulary
       - warns about empty query / reference
  4. Writes eval/gold_set.jsonl (one JSON object per line, UTF-8).
"""

from pathlib import Path
from collections import Counter
import json
import re

from openpyxl import load_workbook

# Paths are built from this file's location, so the script works no matter
# which directory you run it from (same trick as ingest.py).
PROJECT_ROOT = Path(__file__).resolve().parents[1]
XLSX = PROJECT_ROOT / "eval" / "gold_worksheet.xlsx"
TM_PATH = PROJECT_ROOT / "data" / "tm" / "translation_memory.jsonl"
OUT = PROJECT_ROOT / "eval" / "gold_set.jsonl"

# Allowed values - must match the drop-down lists in the worksheet.
EDIT_TYPES = {"name", "term", "number", "gender",
              "addition", "reworded", "omission", "tag/formatting"}
AGREE = {"yes", "no"}
FLAGS = {"ok", "target_error", "needs_client"}

# A Wesnoth variable looks like $name, $unit.name, $student_hp, optionally
# ending with the "|" terminator (e.g. $unit.name|). We use this to check that
# placeholders survive from the query into the reference translation.
# A dot only counts as part of the name when a letter follows it ($unit.name),
# so a sentence-ending period after a variable ($target_name.) is NOT swallowed -
# that ambiguity is exactly why Wesnoth has the "|" terminator.
PLACEHOLDER_RE = re.compile(r"\$[A-Za-z_][A-Za-z0-9_]*(?:\.[A-Za-z_][A-Za-z0-9_]*)*\|?")


def load_tm_ids():
    """Return the set of every record id in the translation memory.
    Used to check that each `base_id` in the edited sheet is real."""
    ids = set()
    with open(TM_PATH, encoding="utf-8") as f:
        for line in f:
            ids.add(json.loads(line)["id"])
    return ids


def clean(value):
    """Trim a cell value; turn blanks into None so empty cells are consistent."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def data_rows(ws):
    """Yield (row_index, {header: value}) for real data rows only.
    Skips row 1 (headers) and row 2 (the illustrative EXAMPLE)."""
    headers = [ws.cell(1, j).value for j in range(1, ws.max_column + 1)]
    for i in range(3, ws.max_row + 1):
        record = {h: ws.cell(i, j).value for j, h in enumerate(headers, start=1)}
        if all(v in (None, "") for v in record.values()):
            continue  # skip fully empty rows
        yield i, record


def placeholders(text):
    """Set of Wesnoth $placeholders found in a piece of text (empty if none)."""
    return set(PLACEHOLDER_RE.findall(text or ""))


def main():
    tm_ids = load_tm_ids()
    wb = load_workbook(XLSX)

    out_records = []
    problems = []  # list of (case_id, message)

    for sheet in ("fuzzy_real", "edited", "invented"):
        ws = wb[sheet]
        for i, row in data_rows(ws):
            cid = clean(row.get("case_id")) or f"{sheet}:row{i}"
            query = clean(row.get("query"))
            reference = clean(row.get("reference"))
            edit_type = clean(row.get("edit_type"))
            req_agree = clean(row.get("requires_agreement"))
            flag = clean(row.get("flag"))

            # Common fields shared by every category.
            rec = {
                "case_id": cid,
                "category": sheet,
                "query": query,
                "reference": reference,
                "edit_type": edit_type,
                "requires_agreement": req_agree,
                "flag": flag,
                "comment": clean(row.get("comment")),
            }

            # Category-specific fields.
            if sheet == "fuzzy_real":
                # Real source change; target_approved is the repair baseline,
                # NOT a reference for the new source.
                rec["previous_source"] = clean(row.get("previous_source"))
                rec["target_approved"] = clean(row.get("target_approved"))
                rec["source_file"] = clean(row.get("source_file"))
            elif sheet == "edited":
                base_id = clean(row.get("base_id"))
                rec["base_id"] = base_id
                rec["source"] = clean(row.get("source"))
                rec["target"] = clean(row.get("target"))
                rec["source_file"] = clean(row.get("source_file"))
                # acceptable_ids = the retrieval ground truth. For now just the
                # edited entry itself. Sibling handling (e.g. Li'sar vs Konrad)
                # is a policy decision for the retrieval eval, so we leave it
                # to that step instead of baking a guess in here.
                rec["acceptable_ids"] = [base_id] if base_id else []

            # ---- validations (reported, never fatal) ----
            if not query:
                problems.append((cid, "query is empty"))
            if not reference:
                problems.append((cid, "reference is empty (still to be written?)"))
            if edit_type and edit_type not in EDIT_TYPES:
                problems.append((cid, f"edit_type '{edit_type}' is not in the allowed list"))
            if req_agree and req_agree not in AGREE:
                problems.append((cid, f"requires_agreement '{req_agree}' is not yes/no"))
            if flag and flag not in FLAGS:
                problems.append((cid, f"flag '{flag}' is not in the allowed list"))
            if sheet == "edited" and rec.get("base_id") and rec["base_id"] not in tm_ids:
                problems.append((cid, f"base_id not found in the translation memory: {rec['base_id']}"))
            # Placeholder preservation: only check when a reference exists.
            if reference:
                lost = placeholders(query) - placeholders(reference)
                if lost:
                    problems.append((cid, f"placeholder(s) missing from reference: {sorted(lost)}"))

            out_records.append(rec)

    # ---- write the JSONL ----
    with open(OUT, "w", encoding="utf-8") as f:
        for rec in out_records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")

    # ---- report ----
    by_cat = Counter(r["category"] for r in out_records)
    print(f"Wrote {len(out_records)} cases to {OUT.relative_to(PROJECT_ROOT)}")
    for cat in ("fuzzy_real", "edited", "invented"):
        print(f"  {cat}: {by_cat.get(cat, 0)}")

    if problems:
        print(f"\n{len(problems)} issue(s) to review (rows were still written):")
        for cid, msg in problems:
            print(f"  [{cid}] {msg}")
    else:
        print("\nNo validation issues found.")


if __name__ == "__main__":
    main()
